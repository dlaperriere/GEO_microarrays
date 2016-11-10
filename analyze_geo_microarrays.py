#!/usr/bin/env python
"""
Differential expression analysis of published microarrays datasets from
the NCBI Gene Expression Omnibus (GEO)

Usage

     python analyze_geo_microarrays.py  -g MCF7_E2_CHX.GSE8597.analysis.txt

geo dataset file format (-g filename, tsv)

parameter | value
--------- | -----------
gse       | GSE#      e.g. GSE8597
gpl       | GPL#      e.g. GPL570
samples   | filename
contrast  | treatment-control    e.g. E2-EtOH
contrast  | treatment2-control2
contrast  | ...

samples file format (tsv)

ID            | sample                          | condition
------------- | ------------------------------- | ---------
GSM213318     | MCF7_CHX_E2_24h_rep1            | CHX_E2
GSM213322     | MCF7_CHX_EtOH_24h_rep1          | CHX_EtOH
GSM213326     | MCF7_E2_24h_rep1                | E2
GSM213330     | MCF7_EtOH_24h_rep1              | EtOH

Copyright

David Laperriere dlaperriere@outlook.com
"""
from __future__ import print_function

import argparse
import csv
import glob
import os
import re
import textwrap

from openpyxl import Workbook
from openpyxl.drawing.image import Image

from utils import r

__version_info__ = (1, 0)
__version__ = '.'.join(map(str, __version_info__))
__author__ = "David Laperriere dlaperriere@outlook.com"


def base_dir():
    """Find python script base directory"""
    return os.path.dirname(os.path.realpath(__file__))


def build_argparser():
    """ Build command line arguments parser"""
    parser = argparse.ArgumentParser(
        prog="analyze_geo_microarrays.py",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent('''
      Differential expression analysis of published microarrays datasets
                from the NCBI Gene Expression Omnibus (GEO)
                               '''),
        epilog=textwrap.dedent('''
      geo dataset file format (tsv)

      gse      | GSE#       |           e.g. GSE8597
      gpl      | GPL#       |           e.g. GPL570
      samples  | filename   |
      contrast | treatment-control      e.g. E2-EtOH
      contrast | treatment2-control2
      contrast | ...

      samples file format (tsv)

      ID            | sample                          | condition
      ------------- | ------------------------------- | ---------
      GSM213318     | MCF7_CHX_E2_24h_rep1            | CHX_E2
      GSM213322     | MCF7_CHX_EtOH_24h_rep1          | CHX_EtOH
      GSM213326     | MCF7_E2_24h_rep1                | E2
      GSM213330     | MCF7_EtOH_24h_rep1              | EtOH
                       ''')
    )
    parser.add_argument('-g', '--geo_dataset',
                        type=argparse.FileType('r'),
                        required=True,
                        help='microarrays information and conditions')
    parser.add_argument('-v', '--version', action='version',
                        version="%(prog)s v" + __version__)

    return parser


def check_analysis_parameters(parameters):
    """ Verify that the provided geo dataset contains all parameters """
    required_parameters = ("gse", "gpl", "samples", "contrast")
    for parameter in required_parameters:
        if parameter not in parameters:
            print(("missing {} section in geo dataset file...".format(parameter)))
            exit(-1)

    if not os.path.isfile(parameters["samples"]):
        print(("cannot read sample file {} ...".format(parameters["samples"])))
        exit(-1)

    gpl_info = os.path.join(base_dir(), "R", "gpl_info.txt")
    with open(gpl_info) as gpl_info_file:
        info = gpl_info_file.read()
        if not re.search("{}\t".format(parameters["gpl"]), info):
            print(("{} missing from gpl_info file {} ...".format(
                parameters["gpl"], gpl_info)))
            exit(-1)


def excel_add_tsv(filename, title, workbook):
    """
    Add content of a tsv file to an excel worksheet

        parameters
         - filename: tsv file name
         - title: worksheet name
         - workbook: openpyxl workbook
    """
    invalid_char = r'[\\*?:/\[\]]'  # http://openpyxl.readthedocs.io/en/2.3.3/_modules/openpyxl/workbook/child.html
    title = re.sub(invalid_char, '_', title)
    max_length = 31
    if len(title) > max_length:
        title = title[0:max_length - 1]

    ws = workbook.create_sheet(title=title)

    with open(filename) as tab_file:
        tab_reader = csv.reader(tab_file, delimiter='\t')
        for idx, line in enumerate(tab_reader):
            for column in range(len(line)):
                _ = ws.cell(row=idx + 1, column=column + 1, value=line[column])


def excel_diffexpression(parameters, excel_file):
    """
    Create excel file with all contrasts
        parameters
         - parameters: analysis parameters dictionary
         - excel_file: excel file name
    """
    wb = Workbook(guess_types=True)
    ws = wb.active

    ws.title = "Dataset"
    ws.cell(column=1, row=1,
            value="http://www.ncbi.nlm.nih.gov/geo/query/acc.cgi?acc={}".format(parameters["gse"]))

    methods = os.path.join(base_dir(), "R", "methods.txt")
    excel_add_tsv(methods, "Methods", wb)

    diffexp_files = glob.glob('DiffExpression/diffexp.*.txt')
    for diffexp_file in diffexp_files:
        print('.', end=' ')
        contrast = os.path.basename(diffexp_file)
        contrast = re.sub(r"diffexp.GSE\w*.", "", contrast)
        contrast = re.sub(r".txt", "", contrast)
        excel_add_tsv(diffexp_file, contrast, wb)

    images_files = glob.glob('Figures/*.png')
    ws = wb.create_sheet(title="Figures")
    linenum = 1
    for image_file in images_files:
        print('.', end=' ')
        img = Image(image_file)
        ws.add_image(img, "A{}".format(linenum))
        linenum += 28

    wb.save(excel_file)


def parse_analysis_parameters(parameters_file):
    """
    Parse analysis parameters file (-g filename)

            **file format (tsv)**
            parameter | value
            --------- | -----------
            gse       | GSE8597
            gpl       | GPL570
            samples   | filename
            contrast  | treatment-control
            contrast  | treatment2-control2
            contrast  | E2-EtOH
    """
    parameters = dict()
    for line in parameters_file:
        column = line.strip().split("\t")
        if len(column) < 2:
            continue
        section = column[0].lower()
        if "contrast" in section:
            if "contrast" not in parameters:
                parameters[section] = list()
            parameters[section].append(column[1])
        else:
            parameters[section] = column[1]
    return parameters


def main():
    """ Run microarray analysis with R """
    # parse command line arguments
    parser = build_argparser()
    pyargs = parser.parse_args()

    # make sure R is available
    Rpath = r.findR()
    if Rpath is None:
        print("Could not find R scripting front-end path (Rscript)")
        exit(-1)
    print(("R path: {}\n".format(Rpath)))

    # parse analysis parameters file (-g filename)
    parameters = parse_analysis_parameters(pyargs.geo_dataset)
    check_analysis_parameters(parameters)

    # run analysis for each contrast
    analysis_script = os.path.join(
        base_dir(), "R", "analyze_geo_microarrays.R")
    gpl_info = os.path.join(base_dir(), "R", "gpl_info.txt")

    for contrast in parameters["contrast"]:
        print(("\nContrast: {}".format(contrast)))
        rargs = "--gse {} --gse_samples {} --gpl {} --gpl_info {} --contrast {} ".format(
            parameters["gse"], parameters["samples"], parameters["gpl"], gpl_info, contrast)
        out, status = r.runR(r=Rpath, script=analysis_script, args=rargs)
        print(out)
        if status == -1:
            exit(-1)

    # create excel file with all contrasts
    excel_file = "DiffExpression.{}.xlsx".format(parameters["gse"])
    print(("\nWriting results to {}".format(excel_file)))
    excel_diffexpression(parameters, excel_file)

    # -30-
    print("\nDone")
    exit(0)

if __name__ == "__main__":
    main()
